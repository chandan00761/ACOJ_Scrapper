from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook


class Scrapper:
    def __init__(self, base_url, *urls):
        self.base_url = base_url
        self.urls = urls
        self.wb = Workbook()

    def scrape(self, url):
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        rows = soup.find_all("table")[1].find_all("tr")[1:]
        for row in rows:
            yield {
                "link": row.contents[3].contents[0]['href'],
                "name": row.contents[3].contents[0].string,
                "site": row.contents[5].contents[0],
                "difficulty": row.contents[7].contents[0],
            }

    def start(self):
        """
        scrapes the base_url to get the various links to different ladder pages
        :return:
        """
        for url in self.urls:
            base_response = requests.get(self.base_url + url)
            soup = BeautifulSoup(base_response.content, 'html.parser')
            ladders = soup.find_all("a")
            ladder_pages = []
            sheets = []
            for ladder in ladders:
                ladder_pages.append(self.base_url + ladder['href'])
                sheets.append(ladder.contents[0])
            for index, page in enumerate(ladder_pages):
                if len(sheets[index]) > 31:
                    sheets[index] = sheets[index].replace("Codeforces Rating", "CFR")
                self.wb.create_sheet(sheets[index])
                sheet = self.wb[sheets[index]]
                sheet['A1'] = "PROBLEM"
                sheet['B1'] = "LINK"
                sheet['C1'] = "SITE"
                sheet['D1'] = "DIFFICULTY"
                i = 2
                for item in self.scrape(page):
                    name_cell = sheet.cell(row=i, column=1)
                    link_cell = sheet.cell(row=i, column=2)
                    site_cell = sheet.cell(row=i, column=3)
                    diff_cell = sheet.cell(row=i, column=4)
                    name_cell.value = item["name"]
                    link_cell.value = item["link"]
                    site_cell.value = item["site"]
                    diff_cell.value = item["difficulty"]
                    i = i + 1
            self.wb.save("document.xlsx")
